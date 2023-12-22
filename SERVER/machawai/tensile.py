# --------------
# --- IMPORT ---
# --------------

import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import glob
from machawai.labels import *
from machawai.const import *

# ------------------
# --- EXCEPTIONS ---
# ------------------

class MissingTimeError(Exception):

    def __init__(self, message="No Time values provided") -> None:
        self.message = message
        super().__init__(self.message)

class MissingExtsError(Exception):

    def __init__(self, message="No Extensometer values provided") -> None:
        self.message = message
        super().__init__(self.message)

# ---------------
# --- CLASSES ---
# ---------------

class LabelledObject():

    def __init__(self) -> None:
        pass

    def labels(self) -> 'list[str]':
        raise NotImplementedError()
    
    def get_by_label(self, label:str):
        raise NotImplementedError()
    
    def get_by_labels(self, labels:'list[str]'):
        raise NotImplementedError()

class TestData(LabelledObject):
    """
    Holds the detected tensile test data.
    """
    def __init__(self, 
                 disp: 'pd.Series',
                 load: 'pd.Series',
                 exts: 'pd.Series' = None,
                 time: 'pd.Series' = None) -> None:
        """
        `TestData` class constructor.

        Arguments:
        ----------

        disp: Series
            Detected displacement.

        load: Series
            Applied load.

        exts: Series | ndarray
            True specimen extension.

        time: Series | ndarray
            Time data.
        """
        self.disp = self.handleSeriesInput(disp, acceptNone=False, name=DISPLACEMENT)
        self.load = self.handleSeriesInput(load, acceptNone=False, name=LOAD)
        self.exts = self.handleSeriesInput(exts, acceptNone=True, name=EXTENSOMETER)
        self.time = self.handleSeriesInput(time, acceptNone=True, name=TIME)

        if self.disp.shape[0] != self.load.shape[0]:
            raise ValueError("All specified data series must have the same length.")
        elif isinstance(self.exts, pd.Series) and self.exts.shape[0] != self.load.shape[0]:
            raise ValueError("All specified data series must have the same length.")
        elif isinstance(self.time, pd.Series) and self.time.shape[0] != self.load.shape[0]:
            raise ValueError("All specified data series must have the same length.")

    def handleSeriesInput(self, data: 'pd.Series | np.ndarray | list', acceptNone: bool = False, name: str = "") -> np.ndarray:
        if isinstance(data, pd.Series):
            if data.empty:
                raise ValueError("Error on " + name + " data: empty sequence.")
            data.name = name
            return data
        if isinstance(data, np.ndarray):
            if not data.any():
                raise ValueError("Error on " + name + " data: empty sequence.")
            if len(data.shape) != 1:
                raise ValueError("Error on " + name + " data: invalid shape.")
            return pd.Series(data, name = name)
        if isinstance(data, list):
            if len(data) <= 0:
                raise ValueError("Error on " + name + " data: empty sequence.")
            return pd.Series(data, name = name)
        try:
            if data == None and acceptNone:
                return data
            raise TypeError("Error on " + name + " data: inavlid type.")
        except:
            raise TypeError("Error on " + name + " data: inavlid type.")

    def hasExts(self) -> bool:
        if isinstance(self.exts, pd.Series):
            return not self.exts.empty
        return False
    
    def hasTime(self) -> bool:
        if isinstance(self.time, pd.Series):
            return not self.time.empty
        return False
    
    def getData(self):
        if self.hasTime():
            if self.hasExts():
                return pd.DataFrame({TIME: self.time,
                                     DISPLACEMENT: self.disp,
                                     LOAD: self.load,
                                     EXTENSOMETER: self.exts})
            return pd.DataFrame({TIME: self.time,
                                 DISPLACEMENT: self.disp,
                                 LOAD: self.load})
        else:
            if self.hasExts():
                return pd.DataFrame({DISPLACEMENT: self.disp,
                                     LOAD: self.load,
                                     EXTENSOMETER: self.exts})
        return pd.DataFrame({DISPLACEMENT: self.disp,
                             LOAD: self.load})

    def labels(self) -> 'list[str]':
        return [TIME, DISPLACEMENT, LOAD, EXTENSOMETER]
    
    def get_by_label(self, label:str):
        return self.getData()[label]
    
    def get_by_labels(self, labels: 'list[str]'):
        return self.getData()[labels]

class SpecimenProperties(LabelledObject):
    """
    Holds the geometric properties of the tested specimen.
    """
    def __init__(self,
                 width: 'float | list',
                 thickness: 'float | list',
                 interaxis: float,
                 constant_section_length: float,
                 exts_length: float = None) -> None:
        """
        `SpecimenProperties` class constructor.

        Arguments:
        ----------

        width: float | list
            Mesured specimen width.

        thickness: float | list
            Mesuerd speciemen thickness.

        interaxis: float
            Interaxis distance.

        exts_length: float
            Initial extensometer length.

        constant_section_length: float
            Constant section length.
        """
        if isinstance(width, float):
            self.width = [width]
        elif isinstance(width, list):
            if len(width) <= 0:
                raise ValueError("Empty width.")
            self.width = width
        else:
            raise TypeError("Inavlid data type for width argument.")
        
        if isinstance(thickness, float):
            self.thickness = [thickness]
        elif isinstance(thickness, list):
            if len(thickness) <= 0:
                raise ValueError("Empty thickness.")
            self.thickness = thickness
        else:
            raise TypeError("Inavlid data type for thickness argument.")       

        self.interaxis = interaxis
        self.exts_length = exts_length
        self.constant_section_length = constant_section_length

        self.trasversal_section = self.meanWidth() * self.meanThickness()

    def meanWidth(self):
        return np.mean(self.width)
    
    def meanThickness(self):
        return np.mean(self.thickness)
    
    def getProperties(self) -> dict:
        return {
            WIDTH: self.width,
            THICKNESS: self.thickness,
            INTERAXIS: self.interaxis,
            EXTS_LENGTH: self.exts_length,
            CS_LENGTH: self.constant_section_length
        }
    
    def labels(self) -> 'list[str]':
        return [WIDTH, THICKNESS, INTERAXIS, EXTS_LENGTH, CS_LENGTH]
    
    def get_by_label(self, label: str):
        return self.getProperties()[label]
    
    def get_by_labels(self, labels: 'list[str]'):
        toReturn = {}
        prop = self.getProperties()
        for label in labels:
            toReturn[label] = prop
        return toReturn

class TestSetup(LabelledObject):
    """
    Additional Tensile Test information.
    """
    def __init__(self,
                 extensometer_acquired: bool = True,
                 ref_displacement_load: str = MOTOR,
                 ref_param_strain: str = EXTENSO,
                 linearity_dev_method: str = RP02,
                 straingage1_acquired: bool = False,
                 straingage2_acquired: bool = False) -> None:
        """
        `TestSetup` class constructor.

        Arguments:
        ----------

        extensometer_acquired: bool
            `True` if an Extensometer has was used during the Tensile Test.

        ref_displacement_load: int
            Source of Displacement data.

        ref_param_strain: int
            Data source for Strain calculation.

        linearity_dev_method: int
            Linearity Deviation method.

        straingage1_acquired: bool
            TBD

        straingage2_acquired: bool
            TBD
        
        """
        if not ref_displacement_load in ACCEPTED_REF_DISPLACEMENT_LOAD:
            raise ValueError("Invalid ref_displacement_load.")
        if not ref_param_strain in ACCEPTED_REF_PARAM_STRAIN:
            raise ValueError("Invalid ref_param_strain.")
        if not linearity_dev_method in LINEARITY_DEV_METHODS:
            raise ValueError("Invalid linearity_dev_method.")
        
        self.extensometer_acquired = extensometer_acquired
        self.straingage1_acquired = straingage1_acquired
        self.straingage2_acquired = straingage2_acquired
        self.ref_displacement_load = ref_displacement_load
        self.ref_param_strain = ref_param_strain
        self.linearity_dev_method = linearity_dev_method

    def getSetup(self) -> dict:
        return {
            EXTS_ACQUIRED: self.extensometer_acquired,
            REF_DISP_LOAD: self.ref_displacement_load,
            REF_PARAM_STRAIN: self.ref_param_strain,
            LINEARITY_DEV_METHOD: self.linearity_dev_method,
            STRAINGAGE1_ACQUIRED: self.straingage1_acquired,
            STRAINGAGE2_ACQUIRED: self.straingage2_acquired
        }
    
    def labels(self) -> 'list[str]':
        return [EXTS_ACQUIRED, REF_DISP_LOAD, REF_PARAM_STRAIN, LINEARITY_DEV_METHOD, STRAINGAGE1_ACQUIRED, STRAINGAGE2_ACQUIRED]
    
    def get_by_label(self, label: str):
        return self.getSetup()[label]
    
    def get_by_labels(self, labels: 'list[str]'):
        toReturn = {}
        prop = self.getSetup()
        for label in labels:
            toReturn[label] = prop
        return toReturn

class CutAndOffset(LabelledObject):

    def __init__(self,
                 tail_p: float = 0.0,
                 foot_offset: float = 0.0) -> None:
        """
        Arguments:
        ----------

        tail_p: float:
            Percentage of data points to exclude at the end of the curve.

        foot_offset: float
            Foot offset to apply to data.
        """
        self.tail_p = tail_p
        self.foot_offset = foot_offset

    def apply(self, curve: pd.DataFrame, col_idx: int = 0):
        curve_length = curve.shape[0]
        to_tail = curve_length * self.tail_p
        curve = curve.iloc[:curve_length - to_tail]

        curve.iloc[:,col_idx] = curve.iloc[:,col_idx] - self.foot_offset
        curve = curve.loc[curve.iloc[:,col_idx] < 0]
        return curve
    
    def getCutAndOffset(self):
        return {
            TAIL_P: self.tail_p,
            FOOT_OFFSET: self.foot_offset
        }
    
    def labels(self) -> 'list[str]':
        return [TAIL_P, FOOT_OFFSET]
    
    def get_by_label(self, label: str):
        return self.getCutAndOffset()[label]
    
    def get_by_labels(self, labels: 'list[str]'):
        toReturn = {}
        prop = self.getCutAndOffset()
        for label in labels:
            toReturn[label] = prop
        return toReturn

class LinearSection(LabelledObject):

    def __init__(self,
                 bottom_cutout: float = 0.0,
                 upper_cutout: float = None) -> None:
        """
        Arguments:
        ----------

        bottom_cutout: float
            Bottom Stress cutout.

        upper_cutout:
            Upper Stress cutout.
        """
        self.bottom_cutout = bottom_cutout
        self.upper_cutout = upper_cutout

    def apply(self, curve: pd.DataFrame, col_idx: int = 1):
        upper_cutout = self.upper_cutout
        if upper_cutout == None:
            upper_cutout = curve.iloc[-1, col_idx]
        curve = curve.loc[curve.iloc[:,col_idx] >= self.bottom_cutout]
        curve = curve.loc[curve.iloc[:,col_idx] <= upper_cutout]

        return curve
    
    def getBounds(self):
        return {
            BOTTOM_CUTOUT: self.bottom_cutout,
            UPPER_CUTOUT: self.upper_cutout
        }
    
    def labels(self) -> 'list[str]':
        return [BOTTOM_CUTOUT, UPPER_CUTOUT]
    
    def get_by_label(self, label: str):
        return self.getBounds()[label]
    
    def get_by_labels(self, labels: 'list[str]'):
        toReturn = {}
        prop = self.getBounds()
        for label in labels:
            toReturn[label] = prop
        return toReturn

class TensileTest(LabelledObject):
    """
    Describe a Tensile Test.
    """
    def __init__(self, 
                 testData: TestData,
                 specimenProperties: SpecimenProperties,
                 testSetup: TestSetup,
                 cutAndOffset: CutAndOffset = None,
                 linearSection: LinearSection = None,
                 filename: str = None) -> None:
        """
        `TensileTest` class constructor.

        Arguments:
        ----------

        testData: `TestData`
            Detected tensile test data.

        specimenProperties: `SpecimenProperties`
            Geometric properties of the tested specimen.

        testSetup: `TestSetup`
            Additional test info.

        cutAndOffset: `CutAndOffset`
            Cut and offset options.

        linearSection: `LinearSection`
            Options for linear section localization.

        filename: str:
            Name of source file.
        """
        self.testData = testData
        self.specimenProperties = specimenProperties
        self.testSetup = testSetup
        self.cutAndOffset = cutAndOffset
        self.linearSection = linearSection
        self.filename = filename

        if self.linearSection == None:
            self.linearSection = LinearSection()

        if self.testSetup.extensometer_acquired and not self.testData.hasExts():
            raise ValueError("Inconsistent setup and data: Extensometer is acquired but no value has been provided.")
    
    def getDispStrain(self):
        """
        Compute the Strain values on Displacement.
        """
        return self.testData.disp / self.specimenProperties.exts_length
    
    def getExtsStrain(self):
        """
        Compute the Strain values on Extensometer.
        """
        if self.testData.hasExts():
            strain = self.testData.exts / self.specimenProperties.exts_length
            strain.name = EXTS_STRAIN
            return strain
        else:
            raise MissingExtsError()
    
    def getStress(self):
        """
        Compute the Stress values.
        """
        stress = self.testData.load / self.specimenProperties.trasversal_section
        stress.name = STRESS
        return stress
    
    def getElabData(self):
        data = pd.DataFrame({DISP_STRAIN: self.getDispStrain(),
                             STRESS: self.getStress()})
        if self.testData.hasExts():
            data[EXTS_STRAIN] = self.getExtsStrain()
        return data

    def getFullData(self):
        """
        Return the full Test Dataframe.
        """
        data = self.testData.getData()
        # Add disp/extx strain
        data[DISP_STRAIN] = self.getDispStrain()
        if self.testData.hasExts():
            data[EXTS_STRAIN] = self.getExtsStrain()
        data[STRESS] = self.getStress()

        return data

    def getLoadDisplacementCurve(self, cut_and_offset: bool = True):
        """
        Return the Load-Displacement curve.
        """
        curve = pd.DataFrame({DISPLACEMENT: self.testData.disp,
                              LOAD: self.testData.load})
        if cut_and_offset and self.cutAndOffset != None:
            curve = self.cutAndOffset.apply(curve=curve, col_idx=0)
        return curve
    
    def getLoadExtsCurve(self, cut_and_offset: bool = True):
        """
        Return the Load-Extensometer curve.

        Raise `MissingExtsError` if no Extensometer value is provided.
        """
        if not self.testData.hasExts():
            raise MissingExtsError()
        curve = pd.DataFrame({EXTENSOMETER: self.testData.exts,
                              LOAD: self.testData.load})
        if cut_and_offset and self.cutAndOffset != None:
            curve = self.cutAndOffset.apply(curve=curve, col_idx=0)
        return curve
    
    def getStrainStressCurve(self, use_displacement: bool = False, cut_and_offset: bool = True):
        """
        Return the Strain-Stress curve.

        Arguments:
        ----------

        use_displacement: bool
            If `True` displacement values are used for Strain calculation.
        """
        stress = self.getStress()
        if use_displacement:
            strain = self.getDispStrain()
            curve = pd.DataFrame({DISP_STRAIN: strain,
                                  STRESS: stress})
        else:
            strain = self.getExtsStrain()
            curve = pd.DataFrame({EXTS_STRAIN: strain,
                                  STRESS: stress})
        if cut_and_offset and self.cutAndOffset != None:
            curve = self.cutAndOffset.apply(curve=curve, col_idx=0)
        return curve
    
    def getDataStats(self, labels: 'list[str]' = None):
        if labels == None:
            return self.getFullData().describe()
        else:
            return self.getFullData[labels].describe()

    def labels(self) -> 'list[str]':
        toReturn = [FILENAME, DISP_STRAIN, EXTS_STRAIN, STRESS]
        toReturn.extend(self.testData.labels())
        toReturn.extend(self.specimenProperties.labels())
        toReturn.extend(self.testSetup.labels())
        toReturn.extend(self.cutAndOffset.labels())
        toReturn.extend(self.linearSection.labels())
        return toReturn
    
    def get_by_label(self, label: str):
        if label in self.testData.labels():
            return self.testData.get_by_label(label)
        if label in self.specimenProperties.labels():
            return self.specimenProperties.get_by_label(label)
        if label in self.testSetup.labels():
            return self.testSetup.get_by_label(label)
        if label in self.cutAndOffset.labels():
            return self.cutAndOffset.get_by_label(label)
        if label in self.linearSection.labels():
            return self.linearSection.get_by_label(label)
        if label == FILENAME:
            return self.filename
        return self.getElabData()[label]
    
    def get_by_labels(self, labels: 'list[str]'):
        toReturn = {}
        for label in labels:
            toReturn[label] = self.get_by_label(label)
        return toReturn

# -----------------
# --- FUNCTIONS ---
# -----------------

def string2bool(string: 'str | int') -> bool:
    if isinstance(string, int):
        return string == 1
    return string.lower() in ["true", "t", "1", "yes", "y", "s"]

def readTensileTest(file: str) -> TensileTest:
    """
    Read a Tensile Test from file.
    """
    # Load the workbook
    wb = load_workbook(filename=file, data_only=True)
    # Get the sheet names
    sheet_names = wb.sheetnames
    # Check that RAW end ELAB sheets exist
    assert RAW in sheet_names and ELAB in sheet_names
    # 1) Read the Test data
    df = pd.read_excel(file, sheet_name=RAW, header=[0,1], dtype=float)
    testData = TestData(disp = df[DISP_COL].iloc[:,0],
                        load = df[LOAD_COL].iloc[:,0],
                        exts = df[EXTS_COL].iloc[:,0] if EXTS_COL in df.columns else None,
                        time = df[TIME_COL].iloc[:,0] if TIME_COL in df.columns else None)
    
    # Read Specimen Properties and options
    prop_sheet = wb[ELAB]
    # Read properties table
    table = prop_sheet[SPROP_START: SPROP_END]

    # 2) Specimen Properties
    width = table[0]
    thickness = table[1]
    interaxis = table[2][0].value
    constant_section_length = table[3][0].value
    exts_length = table[4][0].value
    # Width and Thickness needs to be converted from tuple to list
    width = list(map(lambda c: c.value, width))
    thickness = list(map(lambda c: c.value, thickness))
    sprop = SpecimenProperties(width=width,
                               thickness=thickness,
                               interaxis=interaxis,
                               constant_section_length=constant_section_length,
                               exts_length=exts_length)
    
    # 3) Test Setup
    table = prop_sheet[SETUP_START: SETUP_END]
    ext_acquired = string2bool(table[0][0].value)
    s1_acquired = string2bool(table[1][0].value)
    s2_acquired = string2bool(table[2][0].value)
    rdl = table[3][0].value
    if not (rdl in ACCEPTED_REF_DISPLACEMENT_LOAD):
        raise ValueError("Unknown Ref. Displacement Load. Given value: {}. Accepted: {}".format(rdl, ACCEPTED_REF_DISPLACEMENT_LOAD))
    rps = table[4][0].value
    if not (rps in ACCEPTED_REF_PARAM_STRAIN):
        raise ValueError("Unknown Ref. Parameter for Strain Calculation. Given value: {}. Accepted: {}".format(rps, ACCEPTED_REF_PARAM_STRAIN))
    ldm = table[5][0].value
    if not (ldm in LINEARITY_DEV_METHODS):
        raise ValueError("Unknow Linearity Deviation Method. Given: {}. Accepted: {}".format(ldm, LINEARITY_DEV_METHODS))
    setup = TestSetup(extensometer_acquired = ext_acquired,
                      ref_displacement_load = rdl,
                      ref_param_strain = rps,
                      linearity_dev_method = ldm,
                      straingage1_acquired = s1_acquired,
                      straingage2_acquired = s2_acquired)
    
    # 4) Cut and Offset
    tp = prop_sheet[TAIL_P_CELL].value / 100
    foot_offset = prop_sheet[FOOT_OFFSET_CELL].value / 100
    cao = CutAndOffset(tail_p=tp, foot_offset=foot_offset)

    # 5) Linear Section
    bc = prop_sheet[BOTTOM_CUTOUT_CELL].value
    uc = prop_sheet[UPPER_CUTOUT_CELL].value
    linsec = LinearSection(bottom_cutout=bc, upper_cutout=uc)

    filename = os.path.basename(file)
    filename = os.path.splitext(filename)[0]

    return TensileTest(testData=testData,
                       specimenProperties=sprop,
                       testSetup=setup,
                       cutAndOffset=cao,
                       linearSection=linsec,
                       filename=filename)

def readTensileTestCollection(root: 'str | list[str]', exts: 'list[str]' = ['xlsx']) -> 'list[TensileTest]':
    # TODO: ottimizzare gestione estensioni
    """
    Read a collection of Tensile Tests from a root directory.
    """
    tt_collection = []
    if isinstance(root, str):
        root = [root]
    for dir in root:
        filenames = []
        for ext in exts:
            filenames.extend(glob.glob(dir + "/*." + ext))
        for file in filenames:
            try:
                tt_collection.append(readTensileTest(file=file))
            except:
                print("An error occured while reading file {}".format(file))
    return tt_collection
